""" This module create a model to create a schema, parse data and import data into Weaviate """

import csv
import json
import fnmatch
import openpyxl

from .schema import create_schema_from_model
from .schema import merge_schemas
from .schema import add_classification_to_schema
from .parser import parse_excel
from .parser import parse_csv
from .imports import import_entities
from .imports import import_datapoints
from .imports import set_cross_references
from .classify import execute_classification
from .classify import import_and_classify
from .query import get_classified_datapoints
from .process_result import calculate_validation_score

from .utilities import get_weaviate_client
from .utilities import get_class_name_from_key
from .utilities import get_maxbatch
from .utilities import get_verbose
from .utilities import get_validation_percentage
from .utilities import get_random_selection
from .utilities import calculate_size_training_set
from .utilities import DEFAULT_CSV_DELIMITER

from .exceptions import UnableToGetWeaviateClient
from .exceptions import UnableToOpenModelFile
from .exceptions import UnableToCreateSchema
from .exceptions import UnsupportedModelType
from .exceptions import UnknownModelType
from .exceptions import NoModelLoaded
from .exceptions import UnableToLoadSchema
from .exceptions import NoInstanceLoaded
from .exceptions import NoClassificationSet


def _read_model_json(node: dict, current: dict, model: dict) -> dict:
    #pylint: disable=too-many-branches
    """
    Read the data model file from an json file. Note this is a recursive function

    Parameters
    ----------
    node : dict
        the json node that we are processing in this recursive iteration
    current : dict
        the current class in the model
    model : dict
        The dict containing the model

    Returns
    ------
    dict
        The dict describing the data model
    """

    if model is None:
        model = {}

    for key in node:
        if isinstance(node[key], (list, dict)):
            classname = get_class_name_from_key(key)
            if current is not None:
                current['of'+classname] = classname

            if classname not in model:
                model[classname] = {}

            for item in node[key]:
                if isinstance(node[key], list):
                    _read_model_json(item, model[classname], model)
                else:
                    _read_model_json(node[key], model[classname], model)

        elif isinstance(node[key], str):
            current[key] = 'string'

        elif isinstance(node[key], int):
            current[key] = 'int'

        elif isinstance(node[key], float):
            current[key] = 'number'

        elif isinstance(node[key], bool):
            current[key] = 'boolean'

        else:
            print(key, "something else")

    return model


def _read_model_excel(sheet: openpyxl.worksheet) -> dict:
    """
    Read the data model file from an excel file

    Parameters
    ----------
    sheet : openpyxl.worksheet
        The excel worksheet that contains the model

    Returns
    ------
    dict
        The dict describing the data model
    """

    # Initialize the return value
    model = {}
    model['classes'] = []

    # process the data line by line
    end = False
    row = 1
    while not end:

        if isinstance(sheet.cell(row=row, column=1).value, str):
            newclass = {}
            newclass['classname'] = sheet.cell(row=row, column=1).value
            newclass['columns'] = []
            model['classes'].append(newclass)

        elif isinstance(sheet.cell(row=row, column=1).value, int):
            number = int(sheet.cell(row=row, column=1).value)
            column = {}
            column['number'] = number
            name = sheet.cell(row=row, column=2).value
            if name.startswith("id:"):
                column['name'] = name[3:]
                column['id'] = True
            else:
                column['name'] = name
                column['id'] = False

            column['type'] = sheet.cell(row=row, column=3).value
            column['entity'] = bool(sheet.cell(row=row, column=4).value)
            column['indexInverted'] = bool(sheet.cell(row=row, column=5).value)

            newclass['columns'].append(column)

        elif sheet.cell(row=row, column=1).value is None:
            end = True

        row += 1

    return model


def _read_model(path):

    model = {}
    model['model'] = None
    model['type'] = ""

    if fnmatch.fnmatch(path, "*.xls") or fnmatch.fnmatch(path, "*.xlsx"):
        workbook = openpyxl.load_workbook(path, data_only=True)
        if workbook is not None:
            sheet = workbook.active
            if sheet is not None:
                model['model'] = _read_model_excel(sheet)
                model['type'] = "excel"

    elif fnmatch.fnmatch(path, "*.json"):
        with open(path) as jsonfile:
            root = json.load(jsonfile)
            if root is not None:
                model['model'] = _read_model_json(root, None, None)
                model['type'] = "json"

    elif fnmatch.fnmatch(path, "*.cvs"):
        raise UnsupportedModelType("csv")
    elif fnmatch.fnmatch(path, "*.yaml") or fnmatch.fnmatch(path, "*.yml"):
        raise UnsupportedModelType("yaml")
    else:
        raise UnknownModelType()

    if model['model'] is None or model['model'] == {}:
        raise UnableToOpenModelFile()

    return model


#########################################################################################################
# The class "Model"
#########################################################################################################


class Model:
    #pylint: disable=too-many-public-methods
    """
    Class that holds a data model to create schema, parse data and import data into Weaviate
    """

    #########################################################################################################
    # Initialization functions
    #########################################################################################################

    def __init__(self, path: str, instance: dict=None, classification: dict=None):
        """
        Read the data model file.

        Parameters
        ----------
        path : str
            The path where the data model can be found

        Raises
        ------
        TypeError
            If arguments are of a wrong data type;
        UnknownModelType
            If the modeltype is unknown
        UnsupportedModelType
            If the modeltype is not supported yet
        """

        self.client = None
        self.schema = None
        self.buckets = None

        if not isinstance(path, str):
            raise TypeError("path is expected to be dict but is " + str(type(path)))

        model = _read_model(path)
        self.model = model['model']
        self.type = model['type']

        if instance is not None:
            self.set_instance(instance)
        else:
            self.instance = None

        if classification is not None:
            self.set_classification(classification)
        else:
            self.classification = None


    def set_instance(self, instance):
        """
        Sets the Weaviate instance for this model

        Parameters
        ----------
        instance: dict
            A dict that contains all the weaviate parameters

        Raises
        ------
        TypeError
            If arguments are of a wrong data type;
        UnableToGetWeaviateClient
            if we are unable to get the Weaviate client

        Returns
        -------
        weaviate:client
            the weaviate client indicated to by the argument dict instance
        """

        if not isinstance(instance, dict):
            raise TypeError("instance is expected to be dict but is " + str(type(instance)))

        self.instance = instance
        self.client = get_weaviate_client(instance)
        if self.client is None:
            raise UnableToGetWeaviateClient()

        return self.client


    def get_client(self):
        """
        Returns the Weaviate client indicated by the argument instance

        Parameters
        ----------
        instance: dict
            A dict that contains all the weaviate parameters

        Raises
        ------
        UnableToGetWeaviateClient
            if we are unable to get the Weaviate client

        Returns
        -------
        weaviate:client
            the weaviate client indicated to by the argument dict instance
        """

        if self.client is None:
            raise UnableToGetWeaviateClient()

        return self.client


    #########################################################################################################
    # Functions that have to do with the schema generation - create / load / merge.
    #########################################################################################################


    def create_schema(self, instance: dict=None) -> dict:
        """
        Create a schema from the model stored in self.model

        Raises
        ------
        TypeError
            If arguments are of a wrong data type;
        UnableToOpenCreateSchema
            If schema can not generated from the model

        Returns
        ------
        dict
            The dict describing the Weaviate schema
        """

        if self.model is None:
            raise NoModelLoaded()

        if instance is None:
            if self.instance is not None and 'module_name' in self.instance:
                self.schema = create_schema_from_model(self.model, self.type, modulename=self.instance['module_name'])
            else:
                self.schema = create_schema_from_model(self.model, self.type)
        else:
            if 'module_name' in instance:
                self.schema = create_schema_from_model(self.model, self.type, modulename=instance['module_name'])
            else:
                self.schema = create_schema_from_model(self.model, self.type)

        if self.schema is None or self.schema == {}:
            raise UnableToCreateSchema()

        return self.schema


    def get_schema(self):
        """
        returns the schema currently in model
        """
        return self.schema


    def load_schema(self, schema: dict=None, replace: bool=True):
        """
        Loads the schema into Weaviate

        Parameters
        ----------
        schema: dict
            A dict that contains the schema (optional argument)
        replace: bool (optional argument)
            A boolean that indicates whether the current schema in Weaviate needs to be replaced

        Raises
        ------
        UnableToLoadSchema
            If schema can not be loaded
        """

        if schema is None:
            if self.schema is None:
                raise UnableToLoadSchema()
        else:
            self.schema = schema

        if self.client is not None:
            if replace:
                if self.client.schema.contains():
                    self.client.schema.delete_all()
                self.client.schema.create(self.schema)

            elif not self.client.schema.contains():
                self.client.schema.create(self.schema)


    def merge_schemas(self, schema1: dict, schema2: dict) -> dict:
        #pylint: disable=no-self-use
        """
        Merges two schemas

        Parameters
        ----------
        schema1: dict
            A dict that contains the schema that needs to be merged with schema2
        schema2: dict
            A dict that contains the schema that needs to be merged with schema1

        Raises
        ------
        TypeError
            If arguments are of a wrong data type;
        UnableToOpenCreateSchema
            If schema can not generated from the model

        Returns
        ------
        dict
            The dict describing the combined Weaviate schema
        """

        if not isinstance(schema1, dict):
            raise TypeError("schema is expected to be dict but is " + str(type(schema1)))
        if not isinstance(schema2, dict):
            raise TypeError("schema is expected to be dict but is " + str(type(schema2)))

        schema = merge_schemas(schema1, schema2)

        return schema


    #########################################################################################################
    # Functions that have to do with the data - parsing / import / cross referencing
    #########################################################################################################


    def parse_data(self, path: str, delimiter:str=None) -> dict:
        """
        parses the data from file indicated by path and that fits the argument model

        Parameters
        ----------
        model: dict
            A dict that contains the data model
        path: str
            A string that indicates where the datafile can be found

        Raises
        ------
        TypeError
            If arguments are of a wrong data type;
        NoModelLoaded
            If no model was loaded yet

        Returns
        ------
        dict
            The dict describing the Weaviate schema
        """

        if not isinstance(path, str):
            raise TypeError("path is expected to be str but is " + str(type(path)))
        if self.model is None:
            raise NoModelLoaded()

        # initialize the return value
        data = None

        if fnmatch.fnmatch(path, "*.xls") or fnmatch.fnmatch(path, "*.xlsx"):
            workbook = openpyxl.load_workbook(path, data_only=True)
            if workbook is not None:
                sheet = workbook.active
                if sheet is not None:
                    data = parse_excel(self.model, sheet)

        elif fnmatch.fnmatch(path, "*.csv") or fnmatch.fnmatch(path, "*.txt"):
            with open(path) as csvfile:
                if delimiter is None:
                    reader = csv.reader(csvfile, delimiter=DEFAULT_CSV_DELIMITER)
                else:
                    reader = csv.reader(csvfile, delimiter=delimiter)
                data = parse_csv(self.model, reader)

        return data


    def import_entities(self, entities: dict):
        """
        imports the data into Weaviate according to the argument model

        Parameters
        ----------
        entities: dict
            A dict that contains the entities

        Raises
        ------
        TypeError
            If arguments are of a wrong data type;
        NoModelLoaded
            If no model was loaded yet

        Returns
        ------
        dict
            The dict describing the Weaviate schema
        """

        if not isinstance(entities, dict):
            raise TypeError("entities is expected to be dict but is " + str(type(entities)))
        if self.model is None:
            raise NoModelLoaded()

        maxbatch = get_maxbatch(self.instance)
        verbose = get_verbose(self.instance)
        import_entities(self.client, entities, maxbatch, verbose)


    def import_datapoints(self, datapoints: list):
        """
        imports the data into Weaviate according to the argument model

        Parameters
        ----------
        datapoints: list
            A list that contains the datapoints

        Raises
        ------
        TypeError
            If arguments are of a wrong data type;
        NoModelLoaded
            If no model was loaded yet

        Returns
        ------
        dict
            The dict describing the Weaviate schema
        """

        if not isinstance(datapoints, list):
            raise TypeError("datapoints is expected to be list but is " + str(type(datapoints)))
        if self.model is None:
            raise NoModelLoaded()

        maxbatch = get_maxbatch(self.instance)
        verbose = get_verbose(self.instance)

        import_datapoints(self.client, self.model, datapoints, maxbatch, verbose)


    def set_cross_references(self, datapoints: list, entities: dict):
        """
        Sets the cross references between the datapoints and the entities

        Parameters
        ----------
        datapoints: list
            A list that contains the datapoints
        entities: dict
            A dict that contains the entities

        Raises
        ------
        TypeError
            If arguments are of a wrong data type;
        NoModelLoaded
            If no model was loaded yet
        """

        if not isinstance(datapoints, list):
            raise TypeError("datapoints is expected to be list but is " + str(type(datapoints)))
        if not isinstance(entities, dict):
            raise TypeError("entities is expected to be dict but is " + str(type(entities)))
        if self.model is None:
            raise NoModelLoaded()

        maxbatch = get_maxbatch(self.instance)
        verbose = get_verbose(self.instance)
        set_cross_references(self.client, self.model, datapoints, entities, maxbatch, verbose)


    #########################################################################################################
    # Functions that have to do with classification
    #########################################################################################################


    def set_classification(self, classification):
        """
        Sets the classification parameters

        Parameters
        ----------
        classification: dict
            A dict that contains all the classification parameters

        Raises
        ------
        TypeError
            If arguments are of a wrong data type;
        """

        if not isinstance(classification, dict):
            raise TypeError("classification is expected to be dict but is " + str(type(classification)))

        self.classification = classification


    def get_classification(self):
        """
        returns the classification parameters

        Raises
        ------
        NoClassificationSet
            If no classification has been set
        """

        if self.classification is None:
            raise NoClassificationSet()

        return self.classification


    def import_and_classify(self, datapoints: list):
        """
        imports the argument data points and classifies them. This function is needed because
        the number of data points may exceed the max batch of Weaviate

        Parameters
        ----------
        datapoints: list
            A list that contains the datapoints to be imported and classified
        entities: list
            A list that contains the entities

        Raises
        ------
        TypeError
            If arguments are of a wrong data type;
        NoInstanceLoaded
            If no instance of Weaviate is set
        NoModelLoaded
            If no model is loaded yet
        NoClassificationSet
            If no classification is set
        """

        if self.instance is None:
            raise NoInstanceLoaded()
        if self.model is None:
            raise NoModelLoaded()
        if self.classification is None:
            raise NoClassificationSet()

        if not isinstance(datapoints, list):
            raise TypeError("datapoints is expected to be list but is " + str(type(datapoints)))

        import_and_classify(self.instance, self.model, self.classification, datapoints)


    def get_classified_datapoints(self) -> list:
        """
        Gets all the classified points from Weaviate

        Raises
        ------
        TypeError
            If arguments are of a wrong data type;
        NoModelLoaded
            If no model was loaded yet
        UnableToGetWeaviateClient
            If no client is in the model
        NoClassificationSet
            If no classification is in the model

        Returns
        ------
        list
            The list of all classified datapoints
        """

        if self.model is None:
            raise NoModelLoaded()
        if self.client is None:
            raise UnableToGetWeaviateClient()
        if self.classification is None:
            raise NoClassificationSet()

        maxbatch = get_maxbatch(self.instance)

        return get_classified_datapoints(self.client, self.classification, maxbatch)


    def set_classification_flags(self, datapoints: list, validated: bool=True, preClassified: bool=False):
        #pylint: disable=no-self-use
        """
        sets the validated flag for all datapoints to the argument flag

        Parameters
        ----------
        datapoints: list
            A list of all datapoints
        validated: bool
            A boolean that indicates how the flag should be set
        preClassified: bool
            A boolean that indicates how the flag should be set

        Raises
        ------
        TypeError
            If arguments are of a wrong data type;

        Returns
        ------
        dict
            The dict describing the Weaviate schema
        """

        if not isinstance(datapoints, list):
            raise TypeError("datapoints is expected to be list but is " + str(type(datapoints)))

        for datapoint in datapoints:
            datapoint['validated'] = validated
            datapoint['preClassified'] = preClassified


    def classify(self):
        """
        Classified data points according to the parameters in the argument classification dict

        Raises
        ------
        UnableToGetWeaviateClient
            If client is not set
        NoClassificationSet
            If classification is not set
        """

        if self.client is None:
            raise UnableToGetWeaviateClient()
        if self.classification is None:
            raise NoClassificationSet()

        execute_classification(self.client, self.classification)


    def schema_add_classification(self):
        """
        Add classification classes and properties to schema

        Raises
        ------
        UnableToLoadSchema
            If no schema is in model
        NoClassificationSet
            If no classification is set for model
        """

        if self.schema is None:
            raise UnableToLoadSchema()

        if self.classification is None:
            raise NoClassificationSet()

        if self.instance is not None and 'module_name' in self.instance:
            add_classification_to_schema(self.schema, self.classification, modulename=self.instance['module_name'])
        else:
            add_classification_to_schema(self.schema, self.classification)


    def select_training_data(self, datapoints: list):
        """
        Splits the data in the argument into training data and non training data
        """

        percentage = get_validation_percentage(self.classification)
        random = get_random_selection(self.classification)
        maxbatch = get_maxbatch(self.instance)

        if datapoints is not None:
            size = calculate_size_training_set(datapoints, maxbatch, percentage, random)

            count = total = 0
            for point in datapoints:

                count += 1
                total += 1
                if size['random_selection']:
                    # pick a random number and see if this is control group or training group
                    if random.uniform(0, 100) < size['validation_percentage']:
                        training = False
                    else:
                        training = True
                else:
                    # if count equals the modulus, this is control data
                    if count == size['modulus']:
                        training = False
                        count = 0
                    else:
                        training = True

                if training:
                    point['validated'] = True
                    size['training_size'] += 1
                else:
                    point['validated'] = False
                    size['validation_size'] += 1

            if get_verbose(self.instance):
                print("Total number of datapoints ------------:", size['total'])
                print("Validation percentage -----------------:", size['validation_percentage'])
                print("Random selection of validation sample -:", size['random_selection'])
                print("Number of datapoints in training ------:", size['training_size'])
                print("Number of datapoints in validation ----:", size['validation_size'])


    def calculate_validation_score(self, datapoints: list):
        """
        Calculates the result of a validation run - only used in simple testing

        Raises
        ------
        TypeError
            If arguments are of a wrong data type;
        NoModelLoaded
            If no model was loaded yet
        UnableToGetWeaviateClient
            If no client is in the model
        NoClassificationSet
            If no classification is in the model

        Returns
        ------
        list
            The list of all classified datapoints
        """

        if not isinstance(datapoints, list):
            raise TypeError("datapoints is expected to be list but is " + str(type(datapoints)))
        if self.model is None:
            raise NoModelLoaded()
        if self.client is None:
            raise UnableToGetWeaviateClient()
        if self.classification is None:
            raise NoClassificationSet()

        maxbatch = get_maxbatch(self.instance)

        properties = []
        if 'classify_properties' in self.classification:
            for prop in self.classification['classify_properties']:
                properties.append(prop)

        calculate_validation_score(self.client, self.model, self.classification, datapoints, maxbatch)


    def get_buckets(self) -> dict:
        """
        returns the confidence buckets for the classification

        Raises
        ------
        NoClassificationSet
            If no classification is in the model

        Returns
        ------
        dict
            The dict with the confidence buckets
        """

        if self.classification is None:
            raise NoClassificationSet()

        if self.buckets is not None:
            return self.buckets

        self.buckets = {}
        if 'confidence_buckets' in self.classification:
            intervals = self.classification['confidence_buckets']
        else:
            intervals = [0.0, 1.0]

        length = len(intervals)
        for count in range(length-1):
            self.buckets[count] = {}
            self.buckets[count]['lower'] = intervals[count]
            self.buckets[count]['upper'] = intervals[count+1]
            self.buckets[count]['correct'] = 0
            self.buckets[count]['incorrect'] = 0
            self.buckets[count]['count'] = 0

        return self.buckets


    def get_bucket_id(self, score: float) -> int:
        """
        returns the confidence bucket id for a given score

        Returns
        ------
        int
            The id of the correct confidence bucket
        """

        if self.buckets is None:
            self.get_buckets()

        result = 0
        if score > 0.0:
            for bid in self.buckets:
                if self.buckets[bid]['lower'] < score <= self.buckets[bid]['upper']:
                    result = bid

        return result
